"""
进口涂料自动报价系统 - 主程序
"""
import os
import sys
import json
import re
import smtplib
import sqlite3
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List
from fastapi import FastAPI, HTTPException, Depends, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image, ImageDraw, ImageFont
import io
import requests
from apscheduler.schedulers.background import BackgroundScheduler

from database import get_db, engine, init_db
from models import Product, Order, OrderItem, Expense, EmailConfig, SystemConfig, FeishuConfig

app = FastAPI(title="进口涂料自动报价系统", version="1.0.0")

BASE_DIR = Path(__file__).parent.parent
EXPORTS_DIR = BASE_DIR / "exports"
QUOTES_DIR = EXPORTS_DIR / "quotes"
REPORTS_DIR = EXPORTS_DIR / "reports"
QUOTES_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "frontend")), name="static")


def send_feishu_message(text: str, webhook_url: str = None) -> bool:
    if not webhook_url:
        conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT webhook_url FROM feishu_config WHERE is_active = 1 LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        if not result or not result[0]:
            print("未配置飞书Webhook")
            return False
        webhook_url = result[0]

    try:
        payload = {"msg_type": "text", "content": {"text": text}}
        response = requests.post(webhook_url, json=payload, timeout=10)
        return response.status_code == 200
    except Exception as e:
        print(f"飞书消息发送失败: {e}")
        return False


def scheduled_daily_report():
    from database import SessionLocal
    db = SessionLocal()
    try:
        conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT auto_send_report, report_time, report_recipients
            FROM feishu_config
            WHERE is_active = 1 AND auto_send_report = 1 LIMIT 1
        """)
        config = cursor.fetchone()
        conn.close()

        if not config:
            return

        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        report = generate_daily_report(yesterday)
        filepath = save_report_to_excel(report, "daily")
        filename = f"日报表_{yesterday}.xlsx"
        send_email_report(filename, filepath, yesterday)

        summary = report.get("financial_summary", {})
        profit = summary.get("net_profit", 0)
        sales = summary.get("total_sales", 0)
        message = f"""📊 日报提醒 - {yesterday}
💰 销售总额：¥{sales:,.0f}
💵 净利润：¥{profit:,.0f}
📁 报表已发送到邮箱，请查收"""
        send_feishu_message(message)
    except Exception as e:
        print(f"定时任务执行失败: {e}")
    finally:
        db.close()


class ProductCreate(BaseModel):
    model_code: str
    product_name: str
    series_name: str
    cost_price: float
    default_price: Optional[float] = None


class OrderItemCreate(BaseModel):
    product_model: str
    color_code: str
    quantity: int
    unit_price: float


class OrderCreate(BaseModel):
    customer_name: str
    items: List[OrderItemCreate]
    wood_frame_fee: Optional[float] = 0


class OrderCancel(BaseModel):
    order_no: str


class ExpenseCreate(BaseModel):
    expense_date: str
    category: str
    amount: float
    description: Optional[str] = ""


class FeishuMessage(BaseModel):
    text: str


class EmailConfigCreate(BaseModel):
    smtp_server: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    use_ssl: bool = True
    sender_name: str = "涂料报价系统"
    recipient_email: str


class FeishuConfigCreate(BaseModel):
    webhook_url: Optional[str] = ""
    webhook_secret: Optional[str] = ""
    keyword: str = "报价"
    auto_send_report: bool = False
    report_time: str = "21:00"
    report_recipients: Optional[str] = ""


def generate_order_no():
    today = datetime.now().strftime("%Y%m%d")
    conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
    cursor = conn.cursor()
    cursor.execute("SELECT order_no FROM orders WHERE order_no LIKE ? ORDER BY id DESC LIMIT 1", (f"QT-{today}-%",))
    result = cursor.fetchone()
    conn.close()

    if result:
        last_no = int(result[0].split("-")[-1])
        new_no = f"QT-{today}-{last_no + 1:04d}"
    else:
        new_no = f"QT-{today}-0001"
    return new_no


def parse_feishu_message(text: str) -> dict:
    result = {"customer_name": "", "items": [], "wood_frame_fee": 0}
    customer_match = re.search(r'客户[：:]([^，,]+)', text)
    if customer_match:
        result["customer_name"] = customer_match.group(1).strip()

    wood_match = re.search(r'木架费[：:]?\s*(\d+)', text)
    if wood_match:
        result["wood_frame_fee"] = float(wood_match.group(1))

    product_pattern = r'([A-Za-z0-9]+)\s*(\d+)\s*桶[，,]\s*色号[：:]?([A-Za-z0-9\-]+)[，,]\s*价格(\d+)'
    matches = re.findall(product_pattern, text)

    for match in matches:
        result["items"].append({
            "product_model": match[0],
            "color_code": match[2],
            "quantity": int(match[1]),
            "unit_price": float(match[3])
        })
    return result


def get_product_info(db: Session, model_code: str) -> Optional[Product]:
    return db.query(Product).filter(Product.model_code == model_code).first()


def create_excel_quote(order: Order, items: List[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    normal_font = Font(name="微软雅黑", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = "进口涂料报价单"
    ws['A1'].font = Font(name="微软雅黑", size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A2:G2')
    ws['A2'] = "Premium Paint Quotation"
    ws['A2'].font = Font(name="Arial", size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A4'] = "客户名称："
    ws['A4'].font = Font(name="微软雅黑", size=12, bold=True)
    ws['B4'] = order.customer_name

    ws['E4'] = "订单号："
    ws['E4'].font = Font(name="微软雅黑", size=12, bold=True)
    ws['F4'] = order.order_no

    headers = ["序号", "产品系列", "色号", "型号", "单价(¥)", "数量", "金额(¥)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    row = 7
    total_amount = 0
    for idx, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=idx).alignment = header_align
        ws.cell(row=row, column=2, value=item.get("series_name", "")).font = normal_font
        ws.cell(row=row, column=3, value=item.get("color_code", "-")).alignment = header_align
        ws.cell(row=row, column=4, value=item.get("product_model", "")).alignment = header_align
        ws.cell(row=row, column=5, value=item["unit_price"]).number_format = '#,##0.00'
        ws.cell(row=row, column=5).alignment = header_align
        ws.cell(row=row, column=6, value=item["quantity"]).alignment = header_align
        ws.cell(row=row, column=7, value=item["subtotal"]).number_format = '#,##0.00'
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
        total_amount += item["subtotal"]
        row += 1

    if order.wood_frame_fee > 0:
        ws.cell(row=row, column=1, value=len(items) + 1).alignment = header_align
        ws.cell(row=row, column=2, value="木架费").font = normal_font
        ws.cell(row=row, column=3, value="-").alignment = header_align
        ws.cell(row=row, column=4, value="-").alignment = header_align
        ws.cell(row=row, column=5, value=order.wood_frame_fee).number_format = '#,##0.00'
        ws.cell(row=row, column=5).alignment = header_align
        ws.cell(row=row, column=6, value=1).alignment = header_align
        ws.cell(row=row, column=7, value=order.wood_frame_fee).number_format = '#,##0.00'
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
        total_amount += order.wood_frame_fee
        row += 1

    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="合计").font = Font(name="微软雅黑", size=12, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=7, value=total_amount).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(name="Arial", size=12, bold=True)
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")

    row += 2
    ws.cell(row=row, column=1, value="报价说明：").font = Font(name="微软雅黑", size=11, bold=True)
    row += 1
    ws.cell(row=row, column=1, value="1. 以上颜色经客户确认，不退不换").font = normal_font
    row += 1
    ws.cell(row=row, column=1, value="2. 以上价格不含运费，不含税").font = normal_font

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14

    filename = f"{order.order_no}.xlsx"
    filepath = QUOTES_DIR / filename
    wb.save(filepath)
    return str(filepath)


def create_image_quote(order: Order, items: List[dict]) -> str:
    width = 800
    header_height = 120
    item_height = 45
    footer_height = 100
    total_height = header_height + item_height * (len(items) + 2) + footer_height

    img = Image.new('RGB', (width, total_height), color='white')
    draw = ImageDraw.Draw(img)

    try:
        title_font = ImageFont.truetype("msyh.ttc", 28)
        header_font = ImageFont.truetype("msyh.ttc", 18)
        normal_font = ImageFont.truetype("msyh.ttc", 14)
        small_font = ImageFont.truetype("msyh.ttc", 12)
    except:
        title_font = ImageFont.load_default()
        header_font = ImageFont.load_default()
        normal_font = ImageFont.load_default()
        small_font = ImageFont.load_default()

    draw.rectangle([(0, 0), (width, header_height)], fill='#2C3E50')
    draw.text((width // 2, 30), "进口涂料报价单", fill='white', font=title_font, anchor='mm')
    draw.text((width // 2, 65), "Premium Paint Quotation", fill='#CCCCCC', font=normal_font, anchor='mm')
    draw.text((30, 95), f"客户名称：{order.customer_name}", fill='#333333', font=normal_font)
    draw.text((500, 95), f"订单号：{order.order_no}", fill='#333333', font=normal_font)

    table_top = header_height
    headers = ["序号", "产品系列", "色号", "型号", "单价", "数量", "金额"]
    col_widths = [50, 200, 80, 80, 80, 50, 100]
    x_pos = 0
    draw.rectangle([(0, table_top), (width, table_top + 35)], fill='#34495E')
    for i, header in enumerate(headers):
        draw.text((x_pos + col_widths[i] // 2, table_top + 17), header, fill='white', font=normal_font, anchor='mm')
        x_pos += col_widths[i]

    y_pos = table_top + 35
    total_amount = 0
    for idx, item in enumerate(items):
        bg_color = '#F8F9FA' if idx % 2 == 0 else 'white'
        draw.rectangle([(0, y_pos), (width, y_pos + item_height)], fill=bg_color)
        x_pos = 0
        data = [str(idx + 1), item.get("series_name", "")[:15], item.get("color_code", "-"), item.get("product_model", ""), f"¥{item['unit_price']:.0f}", str(item["quantity"]), f"¥{item['subtotal']:.0f}"]
        for i, text in enumerate(data):
            draw.text((x_pos + col_widths[i] // 2, y_pos + item_height // 2), text, fill='#333333', font=normal_font, anchor='mm')
            x_pos += col_widths[i]
        total_amount += item["subtotal"]
        y_pos += item_height

    if order.wood_frame_fee > 0:
        bg_color = '#F8F9FA' if len(items) % 2 == 0 else 'white'
        draw.rectangle([(0, y_pos), (width, y_pos + item_height)], fill=bg_color)
        draw.text((60, y_pos + item_height // 2), "木架费", fill='#333333', font=normal_font, anchor='mm')
        draw.text((680, y_pos + item_height // 2), f"¥{order.wood_frame_fee:.0f}", fill='#333333', font=normal_font, anchor='mm')
        total_amount += order.wood_frame_fee
        y_pos += item_height

    draw.rectangle([(0, y_pos), (width, y_pos + item_height)], fill='#ECF0F1')
    draw.text((550, y_pos + item_height // 2), "合计：", fill='#333333', font=header_font, anchor='mm')
    draw.text((700, y_pos + item_height // 2), f"¥{total_amount:.0f}", fill='#E74C3C', font=header_font, anchor='mm')

    y_pos += item_height + 20
    draw.text((30, y_pos), "报价说明：", fill='#333333', font=normal_font)
    y_pos += 25
    draw.text((30, y_pos), "1. 以上颜色经客户确认，不退不换", fill='#7F8C8D', font=small_font)
    y_pos += 20
    draw.text((30, y_pos), "2. 以上价格不含运费，不含税", fill='#7F8C8D', font=small_font)

    filename = f"{order.order_no}.png"
    filepath = QUOTES_DIR / filename
    img.save(filepath, 'PNG', quality=95)
    return str(filepath)


def send_email(order_no: str, recipient: str, excel_path: str, image_path: str) -> bool:
    conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM email_config WHERE is_active = 1 LIMIT 1")
    config = cursor.fetchone()
    conn.close()

    if not config:
        print("未配置邮件发送")
        return False

    _, smtp_server, smtp_port, username, password, use_ssl, sender_name, recipient_email, _, _ = config

    try:
        msg = MIMEMultipart('related')
        msg['From'] = f"{sender_name} <{username}>"
        msg['To'] = recipient_email
        msg['Subject'] = f"进口涂料报价单 - {order_no}"

        html_content = f"""<html><body><h2>您好！</h2><p>附件为您本次的涂料报价单，详情请查看附件。</p><p>订单号：<strong>{order_no}</strong></p><p><br>此邮件由系统自动发送，请勿回复。</p></body></html>"""
        msg.attach(MIMEText(html_content, 'html', 'utf-8'))

        with open(excel_path, 'rb') as f:
            part = MIMEText(f.read(), 'base64', 'utf-8')
            part.add_header('Content-Disposition', 'attachment', filename=f'{order_no}.xlsx')
            msg.attach(part)

        with open(image_path, 'rb') as f:
            part = MIMEImage(f.read())
            part.add_header('Content-Disposition', 'attachment', filename=f'{order_no}.png')
            msg.attach(part)

        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            server = smtplib.SMTP(smtp_port)
            server.starttls()
        server.login(username, password)
        server.send_message(msg)
        server.quit()
        print(f"邮件发送成功: {order_no}")
        return True
    except Exception as e:
        print(f"邮件发送失败: {e}")
        return False


def generate_daily_report(date: str = None) -> dict:
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
    cursor = conn.cursor()

    cursor.execute("""
        SELECT o.order_no, o.customer_name, o.wood_frame_fee,
               o.total_sales_amount, o.total_cost_amount, o.profit, o.status
        FROM orders o
        WHERE DATE(o.created_at) = ? AND o.status = 'ACTIVE'
        ORDER BY o.id
    """, (date,))
    sales = cursor.fetchall()

    sales_details = []
    total_sales = 0
    total_cost = 0
    total_profit = 0

    for row in sales:
        order_no, customer, wood_fee, sales_amt, cost_amt, profit, status = row
        cursor.execute("""
            SELECT product_model, series_name, color_code, quantity, unit_price, cost_price, subtotal
            FROM order_items WHERE order_id = (SELECT id FROM orders WHERE order_no = ?)
        """, (order_no,))
        items = cursor.fetchall()

        item_details = []
        for item in items:
            item_details.append({
                "product_model": item[0], "series_name": item[1], "color_code": item[2],
                "quantity": item[3], "unit_price": item[4], "cost_price": item[5], "subtotal": item[6]
            })

        sales_details.append({
            "order_no": order_no, "customer": customer, "items": item_details,
            "wood_frame_fee": wood_fee, "sales_amount": sales_amt, "cost_amount": cost_amt,
            "profit": profit, "status": status
        })
        total_sales += sales_amt
        total_cost += cost_amt
        total_profit += profit

    cursor.execute("SELECT category, SUM(amount) as total FROM expenses WHERE expense_date = ? GROUP BY category", (date,))
    expenses = cursor.fetchall()
    expense_summary = {row[0]: row[1] for row in expenses}
    total_expenses = sum(expense_summary.values())
    net_profit = total_profit - total_expenses

    conn.close()
    return {
        "date": date, "sales_details": sales_details, "expense_summary": expense_summary,
        "financial_summary": {"total_sales": total_sales, "total_cost": total_cost,
        "total_expenses": total_expenses, "gross_profit": total_profit, "net_profit": net_profit}
    }


def save_report_to_excel(report: dict, report_type: str = "daily") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "日报表" if report_type == "daily" else "月报表"
    title = f"日销售报表 - {report['date']}" if report_type == "daily" else f"月销售报表 - {report['year']}年{report['month']}月"

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    normal_font = Font(name="微软雅黑", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(name="微软雅黑", size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    ws.cell(row=row, column=1, value="一、销售明细").font = Font(name="微软雅黑", size=14, bold=True)
    row += 1

    headers = ["订单号", "客户", "产品型号", "系列", "色号", "数量", "销售价", "成本价", "利润", "合计"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row += 1
    for sale in report["sales_details"]:
        ws.cell(row=row, column=1, value=sale["order_no"]).border = thin_border
        ws.cell(row=row, column=2, value=sale["customer"]).border = thin_border
        for item in sale["items"]:
            ws.cell(row=row, column=3, value=item["product_model"]).border = thin_border
            ws.cell(row=row, column=4, value=item["series_name"]).border = thin_border
            ws.cell(row=row, column=5, value=item["color_code"]).border = thin_border
            ws.cell(row=row, column=6, value=item["quantity"]).border = thin_border
            ws.cell(row=row, column=7, value=item["unit_price"]).number_format = '#,##0.00'
            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=8, value=item["cost_price"]).number_format = '#,##0.00'
            ws.cell(row=row, column=8).border = thin_border
            item_profit = (item["unit_price"] - item["cost_price"]) * item["quantity"]
            ws.cell(row=row, column=9, value=item_profit).number_format = '#,##0.00'
            ws.cell(row=row, column=9).border = thin_border
            ws.cell(row=row, column=10, value=item["subtotal"]).number_format = '#,##0.00'
            ws.cell(row=row, column=10).border = thin_border
            row += 1
        if sale["wood_frame_fee"] > 0:
            ws.cell(row=row, column=3, value="木架费").border = thin_border
            ws.cell(row=row, column=10, value=sale["wood_frame_fee"]).number_format = '#,##0.00'
            ws.cell(row=row, column=10).border = thin_border
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="二、费用汇总").font = Font(name="微软雅黑", size=14, bold=True)
    row += 1
    ws.cell(row=row, column=1, value="费用类别").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2, value="金额").fill = header_fill
    ws.cell(row=row, column=2).font = header_font
    row += 1

    for category, amount in report["expense_summary"].items():
        ws.cell(row=row, column=1, value=category).border = thin_border
        ws.cell(row=row, column=2, value=amount).number_format = '#,##0.00'
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="三、财务汇总").font = Font(name="微软雅黑", size=14, bold=True)
    row += 1

    summary = report["financial_summary"]
    summary_data = [("销售汇总", summary["total_sales"]), ("成本汇总", summary["total_cost"]),
                   ("费用汇总", summary["total_expenses"]), ("毛利润", summary["gross_profit"]),
                   ("净利润", summary["net_profit"])]

    ws.cell(row=row, column=1, value="项目").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2, value="金额").fill = header_fill
    ws.cell(row=row, column=2).font = header_font
    row += 1

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=value).number_format = '#,##0.00'
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    if report_type == "daily":
        filename = f"日报表_{report['date']}.xlsx"
    else:
        filename = f"月报表_{report['year']}_{report['month']}.xlsx"

    filepath = REPORTS_DIR / filename
    wb.save(filepath)
    return str(filepath)


@app.get("/")
async def root():
    html_path = BASE_DIR / "frontend" / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text(encoding='utf-8'))
    return HTMLResponse(content="<h1>进口涂料自动报价系统</h1><p>管理后台加载中...</p>")


@app.get("/api/products")
async def get_products(db: Session = Depends(get_db)):
    products = db.query(Product).filter(Product.is_active == True).all()
    return [{"id": p.id, "model_code": p.model_code, "product_name": p.product_name,
             "series_name": p.series_name, "cost_price": p.cost_price, "default_price": p.default_price} for p in products]


@app.post("/api/products")
async def create_product(product: ProductCreate, db: Session = Depends(get_db)):
    existing = db.query(Product).filter(Product.model_code == product.model_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="产品型号已存在")
    db_product = Product(**product.dict())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return {"id": db_product.id, "message": "产品添加成功"}


@app.delete("/api/products/{product_id}")
async def delete_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    product.is_active = False
    db.commit()
    return {"message": "产品删除成功"}


@app.post("/api/orders")
async def create_order(order: OrderCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    order_no = generate_order_no()
    total_sales = 0
    total_cost = 0
    items_data = []

    for item in order.items:
        product = get_product_info(db, item.product_model)
        if not product:
            raise HTTPException(status_code=400, detail=f"产品型号 {item.product_model} 不存在")

        subtotal = item.unit_price * item.quantity
        cost = product.cost_price * item.quantity

        items_data.append({
            "product_model": item.product_model, "product_name": product.product_name,
            "series_name": product.series_name, "color_code": item.color_code,
            "quantity": item.quantity, "unit_price": item.unit_price,
            "cost_price": product.cost_price, "subtotal": subtotal
        })
        total_sales += subtotal
        total_cost += cost

    total_sales += order.wood_frame_fee
    profit = total_sales - total_cost

    db_order = Order(order_no=order_no, customer_name=order.customer_name, wood_frame_fee=order.wood_frame_fee,
                     total_sales_amount=total_sales, total_cost_amount=total_cost, profit=profit, status="ACTIVE")
    db.add(db_order)
    db.commit()
    db.refresh(db_order)

    for item_data in items_data:
        db_item = OrderItem(order_id=db_order.id, **item_data)
        db.add(db_item)
    db.commit()

    excel_path = create_excel_quote(db_order, items_data)
    image_path = create_image_quote(db_order, items_data)

    def send_email_task():
        send_email(order_no, "", excel_path, image_path)
    background_tasks.add_task(send_email_task)

    return {"order_no": order_no, "total_amount": total_sales, "profit": profit, "message": "订单创建成功"}


@app.post("/api/orders/cancel")
async def cancel_order(data: OrderCancel, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.order_no == data.order_no).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    if order.status == "CANCELLED":
        raise HTTPException(status_code=400, detail="订单已取消")
    order.status = "CANCELLED"
    order.cancelled_at = datetime.now()
    order.profit = -order.profit
    db.commit()
    return {"message": f"订单 {data.order_no} 已取消"}


@app.get("/api/orders")
async def get_orders(status: str = None, db: Session = Depends(get_db)):
    query = db.query(Order)
    if status:
        query = query.filter(Order.status == status)
    orders = query.order_by(Order.created_at.desc()).all()
    return [{"order_no": o.order_no, "customer_name": o.customer_name, "total_sales_amount": o.total_sales_amount,
             "total_cost_amount": o.total_cost_amount, "profit": o.profit, "status": o.status,
             "created_at": o.created_at.strftime("%Y-%m-%d %H:%M:%S")} for o in orders]


@app.post("/api/feishu/webhook")
async def handle_feishu_webhook(request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    try:
        body = await request.json()
    except:
        body = {}
    text = body.get("text", "")
    event_type = body.get("type", "")
    if event_type == "event_callback":
        event = body.get("event", {})
        text = event.get("message", {}).get("text", "")

    parsed = parse_feishu_message(text)
    if not parsed["customer_name"]:
        return {"error": "无法解析客户名称"}
    if not parsed["items"]:
        return {"error": "无法解析产品信息"}

    for item in parsed["items"]:
        product = get_product_info(db, item["product_model"])
        if not product:
            return {"error": f"产品型号 {item['product_model']} 不存在"}

    order_data = OrderCreate(customer_name=parsed["customer_name"],
                             items=[OrderItemCreate(**item) for item in parsed["items"]],
                             wood_frame_fee=parsed["wood_frame_fee"])
    result = await create_order(order_data, background_tasks, db)

    return {"msg_type": "text", "content": {"text": f"✅ 报价单已生成！\n订单号：{result['order_no']}\n客户：{parsed['customer_name']}\n金额：¥{result['total_amount']:.0f}\n利润：¥{result['profit']:.0f}"}}


@app.post("/api/feishu")
async def handle_feishu_message(message: FeishuMessage, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    parsed = parse_feishu_message(message.text)
    if not parsed["customer_name"]:
        return {"error": "无法解析客户名称"}
    if not parsed["items"]:
        return {"error": "无法解析产品信息"}

    for item in parsed["items"]:
        product = get_product_info(db, item["product_model"])
        if not product:
            return {"error": f"产品型号 {item['product_model']} 不存在"}

    order_data = OrderCreate(customer_name=parsed["customer_name"],
                             items=[OrderItemCreate(**item) for item in parsed["items"]],
                             wood_frame_fee=parsed["wood_frame_fee"])
    result = await create_order(order_data, background_tasks, db)
    return {"message": "报价单已生成", "order_no": result["order_no"]}


@app.post("/api/expenses")
async def create_expense(expense: ExpenseCreate, db: Session = Depends(get_db)):
    db_expense = Expense(**expense.dict())
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    return {"id": db_expense.id, "message": "费用记录成功"}


@app.get("/api/expenses")
async def get_expenses(date: str = None, db: Session = Depends(get_db)):
    query = db.query(Expense)
    if date:
        query = query.filter(Expense.expense_date == date)
    expenses = query.order_by(Expense.created_at.desc()).all()
    return [{"id": e.id, "expense_date": e.expense_date, "category": e.category, "amount": e.amount, "description": e.description} for e in expenses]


@app.get("/api/reports/daily")
async def get_daily_report(date: str = None, db: Session = Depends(get_db)):
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    return generate_daily_report(date)


@app.get("/api/reports/export/daily")
async def export_daily_report(date: str = None, db: Session = Depends(get_db)):
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    report = generate_daily_report(date)
    filepath = save_report_to_excel(report, "daily")
    return {"filepath": filepath, "filename": Path(filepath).name}


@app.post("/api/email/config")
async def setup_email(config: EmailConfigCreate, db: Session = Depends(get_db)):
    db.query(EmailConfig).update({"is_active": False})
    db_config = EmailConfig(**config.dict(), is_active=True)
    db.add(db_config)
    db.commit()
    db.refresh(db_config)
    return {"message": "邮件配置成功"}


@app.get("/api/email/config")
async def get_email_config(db: Session = Depends(get_db)):
    config = db.query(EmailConfig).filter(EmailConfig.is_active == True).first()
    if config:
        return {"smtp_server": config.smtp_server, "smtp_port": config.smtp_port,
                "smtp_username": config.smtp_username, "sender_name": config.sender_name,
                "recipient_email": config.recipient_email}
    return None


@app.post("/api/feishu/config")
async def setup_feishu(config: FeishuConfigCreate, db: Session = Depends(get_db)):
    existing = db.query(FeishuConfig).first()
    if existing:
        existing.webhook_url = config.webhook_url
        existing.keyword = config.keyword
        existing.auto_send_report = config.auto_send_report
        existing.report_time = config.report_time
        existing.report_recipients = config.report_recipients
        existing.is_active = True
    else:
        db_config = FeishuConfig(**config.dict(), is_active=True)
        db.add(db_config)
    db.commit()
    return {"message": "飞书配置保存成功"}


@app.get("/api/feishu/config")
async def get_feishu_config(db: Session = Depends(get_db)):
    config = db.query(FeishuConfig).filter(FeishuConfig.is_active == True).first()
    if config:
        return {"webhook_url": config.webhook_url, "keyword": config.keyword,
                "auto_send_report": config.auto_send_report, "report_time": config.report_time,
                "report_recipients": config.report_recipients}
    return None


def send_email_report(filename: str, filepath: str, date: str):
    conn = sqlite3.connect(str(BASE_DIR / "data" / "paint.db"))
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM email_config WHERE is_active = 1 LIMIT 1")
    config = cursor.fetchone()
    conn.close()

    if not config:
        return

    _, smtp_server, smtp_port, username, password, use_ssl, sender_name, recipient_email, _, _ = config

    try:
        msg = MIMEMultipart()
        msg['From'] = f"{sender_name} <{username}>"
        msg['To'] = recipient_email
        msg['Subject'] = f"涂料日报表 - {date}"
        html_content = f"<html><body><h2>您好！</h2><p>附件为 {date} 的日报表，请查收。</p></body></html>"
        msg.attach(MIMEText(html_content, 'html', 'utf-8'))

        with open(filepath, 'rb') as f:
            part = MIMEText(f.read(), 'base64', 'utf-8')
            part.add_header('Content-Disposition', 'attachment', filename=filename)
            msg.attach(part)

        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            server = smtplib.SMTP(smtp_port)
            server.starttls()
        server.login(username, password)
        server.send_message(msg)
        server.quit()
    except Exception as e:
        print(f"报表发送失败: {e}")


scheduler = BackgroundScheduler()

@app.on_event("startup")
async def startup_event():
    init_db()
    print("系统启动完成!")

    try:
        scheduler.add_job(scheduled_daily_report, 'cron', hour=21, minute=0, id='daily_report')
        scheduler.start()
        print("定时任务已启动：每天21:00自动发送日报表")
    except Exception as e:
        print(f"定时任务启动失败: {e}")


@app.on_event("shutdown")
async def shutdown_event():
    scheduler.shutdown()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
